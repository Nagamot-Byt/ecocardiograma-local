"""Tests unitarios para DataLoader."""
import pytest
import pandas as pd
from src.core.data_loader import DataLoader
from src.models.patient import Patient, Sexo


@pytest.fixture
def temp_ase_files(tmp_path):
    """Crea archivos Excel de referencia temporales para tests."""
    data_hombres = [
        {"Parametro": "DDI (mm)", "Limite_Inferior": 42, "Limite_Superior": 58, "Unidad": "mm"},
        {"Parametro": "FEVI (%)", "Limite_Inferior": 52, "Limite_Superior": 72, "Unidad": "%"},
        {"Parametro": "PSAP (mmHg)", "Limite_Inferior": None, "Limite_Superior": 35, "Unidad": "mmHg"},
    ]
    data_mujeres = [
        {"Parametro": "DDI (mm)", "Limite_Inferior": 38, "Limite_Superior": 52, "Unidad": "mm"},
        {"Parametro": "FEVI (%)", "Limite_Inferior": 54, "Limite_Superior": 74, "Unidad": "%"},
    ]

    path_h = str(tmp_path / "hombres.xlsx")
    path_m = str(tmp_path / "mujeres.xlsx")

    pd.DataFrame(data_hombres).to_excel(path_h, index=False)
    pd.DataFrame(data_mujeres).to_excel(path_m, index=False)

    return path_h, path_m


class TestDataLoader:
    def test_load_references_success(self, temp_ase_files):
        path_h, path_m = temp_ase_files
        loader = DataLoader(path_h, path_m)
        assert loader.load_references() is True

        # Verificar rangos cargados
        rangos_m = loader.reference_ranges.get_all_ranges(Sexo.MASCULINO)
        rangos_f = loader.reference_ranges.get_all_ranges(Sexo.FEMENINO)
        assert "DDI (mm)" in rangos_m
        assert "DDI (mm)" in rangos_f
        assert len(rangos_m) == 3
        assert len(rangos_f) == 2

    def test_load_references_file_not_found(self):
        loader = DataLoader("/nonexistent/h.xlsx", "/nonexistent/m.xlsx")
        assert loader.load_references() is False

    def test_load_patient_from_file(self, temp_ase_files, tmp_path):
        # Crear archivo de datos del paciente
        data = {
            "DDI": [50],
            "FEVI": [60],
            "PSAP": [30],
        }
        filepath = str(tmp_path / "patient_data.csv")
        pd.DataFrame(data).to_csv(filepath, index=False)

        loader = DataLoader(*temp_ase_files)
        patient = Patient()
        result = loader.load_patient_from_file(filepath, patient)

        assert result is True
        assert patient.ddi == 50.0
        assert patient.fevi == 60.0
        assert patient.psap == 30.0

    def test_load_patient_file_not_found(self, temp_ase_files):
        loader = DataLoader(*temp_ase_files)
        patient = Patient()
        result = loader.load_patient_from_file("/nonexistent/file.csv", patient)
        assert result is False

    def test_load_patient_selecciona_fila(self, temp_ase_files, tmp_path):
        """Con varias filas, ``row`` elige cual usar."""
        data = {
            "DDI": [50, 60],
            "FEVI": [60, 30],
        }
        filepath = str(tmp_path / "multi.csv")
        pd.DataFrame(data).to_csv(filepath, index=False)

        loader = DataLoader(*temp_ase_files)

        patient0 = Patient()
        assert loader.load_patient_from_file(filepath, patient0, row=0) is True
        assert patient0.ddi == 50.0
        assert loader.last_loaded_row == 0
        assert loader.last_row_count == 2

        patient1 = Patient()
        assert loader.load_patient_from_file(filepath, patient1, row=1) is True
        assert patient1.ddi == 60.0
        assert loader.last_loaded_row == 1

    def test_load_patient_fila_fuera_de_rango_usa_primera(self, temp_ase_files, tmp_path):
        data = {"DDI": [50, 60], "FEVI": [60, 30]}
        filepath = str(tmp_path / "multi.csv")
        pd.DataFrame(data).to_csv(filepath, index=False)

        loader = DataLoader(*temp_ase_files)
        patient = Patient()
        assert loader.load_patient_from_file(filepath, patient, row=99) is True
        assert patient.ddi == 50.0
        assert loader.last_loaded_row == 0

    def test_template_columns_tienen_mapeo_valido(self):
        from src.core.data_loader import DataLoader

        columnas = DataLoader.template_columns()
        mapeo = DataLoader._build_column_mapping()
        assert len(columnas) >= 20
        for col in columnas:
            assert col in mapeo, f"Columna de plantilla sin mapeo: {col}"

    def test_template_xlsx_generado_se_puede_cargar(self, tmp_path):
        """La plantilla generada con openpyxl debe ser legible por DataLoader."""
        from openpyxl import Workbook

        filepath = str(tmp_path / "plantilla.xlsx")
        wb = Workbook()
        ws = wb.active
        columnas = DataLoader.template_columns()
        for col, (nombre, ejemplo) in enumerate(columnas.items(), start=1):
            ws.cell(row=1, column=col, value=nombre)
            ws.cell(row=2, column=col, value=ejemplo)
        wb.save(filepath)

        loader = DataLoader("/nonexistent/h.xlsx", "/nonexistent/m.xlsx")
        patient = Patient()
        assert loader.load_patient_from_file(filepath, patient) is True
        assert patient.fevi == 60.0
        assert patient.psap == 28.0


class TestReferenceRanges:
    def test_validate_normal_value(self, temp_ase_files):
        path_h, path_m = temp_ase_files
        loader = DataLoader(path_h, path_m)
        loader.load_references()

        result = loader.reference_ranges.validate_value(Sexo.MASCULINO, "DDI (mm)", 50.0)
        assert result["normal"] is True

    def test_validate_low_value(self, temp_ase_files):
        path_h, path_m = temp_ase_files
        loader = DataLoader(path_h, path_m)
        loader.load_references()

        result = loader.reference_ranges.validate_value(Sexo.MASCULINO, "DDI (mm)", 40.0)
        assert result["normal"] is False
        assert "Bajo" in result["mensaje"]

    def test_validate_high_value(self, temp_ase_files):
        path_h, path_m = temp_ase_files
        loader = DataLoader(path_h, path_m)
        loader.load_references()

        result = loader.reference_ranges.validate_value(Sexo.MASCULINO, "DDI (mm)", 60.0)
        assert result["normal"] is False
        assert "Elevado" in result["mensaje"]

    def test_validate_missing_reference(self, temp_ase_files):
        path_h, path_m = temp_ase_files
        loader = DataLoader(path_h, path_m)
        loader.load_references()

        result = loader.reference_ranges.validate_value(Sexo.MASCULINO, "Parametro Inexistente", 50.0)
        assert result["normal"] is None
