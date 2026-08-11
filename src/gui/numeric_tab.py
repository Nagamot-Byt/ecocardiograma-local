"""
Pestana de datos numericos.
Tabla editable donde se ingresan/cargan los valores del ecocardiograma.
"""
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QFormLayout,
    QLabel, QPushButton, QFileDialog,
    QMessageBox, QInputDialog,
    QDoubleSpinBox,
)
from PyQt6.QtCore import pyqtSignal
import re
from typing import Dict, Optional
from src.models.param_registry import build_campos
from src.models.patient import Patient
from src.utils.helpers import strip_accents


class NumericTab(QWidget):
    """Pestana para la entrada y edicion de datos numericos del ecocardiograma."""

    data_changed = pyqtSignal()  # Emitido cuando se modifica algun dato

    # Definicion de campos numericos (attr, label, unidad, min, max, decimals)
    # Fuente unica de verdad: src.models.param_registry
    CAMPOS = build_campos()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.spinbox_map: Dict[str, QDoubleSpinBox] = {}
        self._last_row_count = 0
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)

        # Toolbar superior
        toolbar = QHBoxLayout()
        self.btn_cargar = QPushButton("Cargar desde Archivo (.xlsx/.csv)")
        self.btn_cargar.clicked.connect(self._on_load_file)
        toolbar.addWidget(self.btn_cargar)

        self.btn_limpiar = QPushButton("Limpiar Datos")
        self.btn_limpiar.clicked.connect(self._on_clear)
        toolbar.addWidget(self.btn_limpiar)

        self.btn_plantilla = QPushButton("Plantilla .xlsx")
        self.btn_plantilla.setToolTip(
            "Descarga una plantilla de ejemplo con las columnas esperadas."
        )
        self.btn_plantilla.clicked.connect(self._on_download_template)
        toolbar.addWidget(self.btn_plantilla)

        self.btn_datos_prueba = QPushButton("Datos de Prueba")
        self.btn_datos_prueba.clicked.connect(self._on_test_data)
        toolbar.addWidget(self.btn_datos_prueba)
        toolbar.addStretch()
        layout.addLayout(toolbar)

        # Formulario con SpinBoxes organizado en columnas
        form_widget = QWidget()
        form_layout = QHBoxLayout(form_widget)

        # Columna izquierda (geometria + volumenes)
        col_left = QFormLayout()
        col_left.setSpacing(4)
        # Columna derecha (valvulas + presiones)
        col_right = QFormLayout()
        col_right.setSpacing(4)

        for i, (attr, label, unidad, min_val, max_val, decimals) in enumerate(self.CAMPOS):
            row_layout = QHBoxLayout()
            spin = QDoubleSpinBox()
            spin.setRange(min_val, max_val)
            spin.setDecimals(decimals)
            spin.setSuffix(f" {unidad}")
            spin.setSpecialValueText("-")
            spin.setMinimumWidth(120)
            spin.valueChanged.connect(lambda val, a=attr: self._on_value_changed(a))
            row_layout.addWidget(spin)
            row_layout.addStretch()
            self.spinbox_map[attr] = spin

            lbl = QLabel(label)
            lbl.setStyleSheet("font-size: 10pt;")
            if i < len(self.CAMPOS) // 2:
                col_left.addRow(lbl, row_layout)
            else:
                col_right.addRow(lbl, row_layout)

        form_layout.addLayout(col_left)
        form_layout.addLayout(col_right)
        layout.addWidget(form_widget)
        layout.addStretch()

    def _on_value_changed(self, attr: str):
        self.data_changed.emit()

    def _on_load_file(self):
        filepath, _ = QFileDialog.getOpenFileName(
            self, "Cargar Datos Numericos", "",
            "Archivos Excel (*.xlsx);;Archivos CSV (*.csv);;Todos (*.*)"
        )
        if filepath:
            # Copiar al directorio user_input para trabajar sobre una copia local
            from src.utils.config import load_config
            from src.utils.helpers import copy_to_user_input

            cfg = load_config()
            try:
                dest = copy_to_user_input(filepath, cfg.user_input_dir)
            except OSError:
                dest = filepath
            self.load_from_file(dest)

    def _on_download_template(self):
        """Genera y guarda una plantilla .xlsx con las columnas esperadas."""
        from src.core.data_loader import DataLoader

        filepath, _ = QFileDialog.getSaveFileName(
            self, "Guardar Plantilla", "plantilla_ecocardiograma.xlsx",
            "Archivo Excel (*.xlsx)"
        )
        if not filepath:
            return
        try:
            import openpyxl

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Datos"
            columnas = DataLoader.template_columns()
            for col, nombre in enumerate(columnas, start=1):
                ws.cell(row=1, column=col, value=nombre)
                ws.cell(row=2, column=col, value=columnas[nombre])
            wb.save(filepath)
            QMessageBox.information(
                self, "Plantilla generada",
                "Plantilla guardada correctamente.\n"
                "Complete la primera fila con los valores del paciente y "
                "cargue el archivo en la aplicacion.\n\n"
                "Puede agregar mas filas: al cargar, la aplicacion le "
                "permitira elegir cual usar.",
            )
        except Exception as e:
            QMessageBox.critical(
                self, "Error", f"No se pudo generar la plantilla:\n{e}"
            )

    def _on_clear(self):
        for attr, spin in self.spinbox_map.items():
            spin.setValue(spin.minimum())
        self.data_changed.emit()

    def _on_test_data(self):
        """Llena con datos de prueba para demostracion."""
        test_values = {
            "ddi": 52, "dsi": 35, "ppvi": 11, "sivi": 12,
            "masa_vi": 200, "masa_vi_ind": 110,
            "rvdi": 130, "rvsi": 50, "fevi": 55,
            "diametro_ai": 42, "volumen_ai": 36,
            "diametro_vd": 38, "tad": 19, "fsr": 40,
            "gradiente_media_mi": 3, "gradiente_max_mi": 8,
            "area_mi": 4.5,
            "gradiente_media_ao": 8, "gradiente_max_ao": 25,
            "area_ao": 2.8, "velocidad_insuf_ao": 2.6,
            "psap": 38,
        }
        for attr, val in test_values.items():
            if attr in self.spinbox_map:
                spin = self.spinbox_map[attr]
                spin.setValue(val)
        self.data_changed.emit()

    def load_from_file(self, filepath: str, row: int = 0) -> bool:
        """Carga datos desde un archivo Excel/CSV usando el DataLoader.

        Si el archivo tiene varias filas, permite elegir cual usar (salvo
        que se indique ``row`` explicitamente).
        """
        from src.core.data_loader import DataLoader
        from src.utils.config import load_config

        cfg = load_config()
        loader = DataLoader(cfg.hombres_file, cfg.mujeres_file)
        patient = Patient()
        success = loader.load_patient_from_file(filepath, patient, row=row)
        self._last_row_count = loader.last_row_count

        if success:
            self.populate_from_patient(patient)
            if row == 0 and self._last_row_count > 1:
                self._ask_row_selection(filepath)
        return success

    def _ask_row_selection(self, filepath: str):
        """Si el archivo tiene varias filas, pregunta cual cargar."""
        items = [f"Fila {i + 1}" for i in range(self._last_row_count)]
        choice, ok = QInputDialog.getItem(
            self, "Multiples filas",
            f"El archivo tiene {self._last_row_count} filas.\n"
            "Seleccione la fila que desea cargar:",
            items, 0, False,
        )
        if not ok:
            return
        idx = items.index(choice)
        if idx != 0:
            self.load_from_file(filepath, row=idx)

    def populate_from_patient(self, patient: Patient):
        """Llena los spinbox con los datos de un objeto Patient."""
        for attr, spin in self.spinbox_map.items():
            val = getattr(patient, attr, None)
            if val is not None:
                spin.setValue(float(val))
        self.data_changed.emit()

    def get_patient_data(self) -> Dict[str, Optional[float]]:
        """Retorna un diccionario con los valores ingresados."""
        data = {}
        for attr, spin in self.spinbox_map.items():
            val = spin.value()
            if val > spin.minimum():
                data[attr] = val
            else:
                data[attr] = None
        return data

    def apply_to_patient(self, patient: Patient):
        """Aplica los valores de la interfaz a un objeto Patient."""
        data = self.get_patient_data()
        for attr, val in data.items():
            setattr(patient, attr, val)

    def highlight_validation(self, validation_results: Dict[str, dict]):
        """
        Colorea los campos segun la validacion:
        Verde = normal, Rojo = fuera de rango.
        """
        for attr, spin in self.spinbox_map.items():
            result = self._find_validation_result(attr, validation_results)
            if result is None:
                spin.setStyleSheet("")
                continue
            if result["normal"] is True:
                spin.setStyleSheet(
                    "QDoubleSpinBox { background-color: #d5f5e3; }"
                )
            elif result["normal"] is False:
                spin.setStyleSheet(
                    "QDoubleSpinBox { background-color: #fadbd8; font-weight: bold; }"
                )
            else:
                spin.setStyleSheet("")

    def _find_validation_result(
        self, attr: str, validation_results: Dict[str, dict]
    ) -> Optional[dict]:
        """Busca el resultado de validacion para un atributo.

        Primero intenta un emparejamiento exacto de tokens canonicos y, si no
        hay, cae a emparejamiento por subconjunto (p. ej. "volumen_ai" frente
        a "Volumen AI ind. (ml/m2)").
        """
        for param_name, result in validation_results.items():
            if self._attr_matches_param(attr, param_name, exact=True):
                return result
        for param_name, result in validation_results.items():
            if self._attr_matches_param(attr, param_name, exact=False):
                return result
        return None

    @staticmethod
    def _canonical_tokens(text: str) -> set:
        """Normaliza un nombre de campo a tokens canonicos comparables."""
        norm = strip_accents(text).replace("_", " ")
        norm = re.sub(r"\(.*?\)", " ", norm)      # quita unidades entre parentesis
        norm = re.sub(r"[^\w\s]", " ", norm)      # puntos, barras -> espacio
        tokens = norm.split()
        return {NumericTab._SINONIMOS.get(t, t) for t in tokens if t}

    # Abreviaturas/sinonimos que deben unificarse al comparar campos
    _SINONIMOS = {
        "grad": "gradiente",
        "media": "medio",
        "max": "maximo",
        "vel": "velocidad",
        "insuf": "insuficiencia",
        "diam": "diametro",
        "ind": "indexada",
        "tad": "tapse",
    }

    @staticmethod
    def _attr_matches_param(attr: str, param_name: str, exact: bool = False) -> bool:
        """Verifica si un atributo corresponde a un nombre de parametro.

        Con ``exact=True`` exige el mismo conjunto de tokens (preferido).
        Con ``exact=False`` acepta que el atributo sea subconjunto del nombre
        (cubre etiquetas con mas palabras, p. ej. "Volumen AI Indexado").
        """
        a = NumericTab._canonical_tokens(attr)
        p = NumericTab._canonical_tokens(param_name)
        if not a or not p:
            return False
        if exact:
            return a == p
        return a <= p

    def clear_highlights(self):
        """Limpia los colores de validacion."""
        for spin in self.spinbox_map.values():
            spin.setStyleSheet("")
